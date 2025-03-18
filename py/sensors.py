from glob import glob
from openpyxl import load_workbook
import json
import requests
import os
from concurrent.futures import ThreadPoolExecutor
import random
import shutil


def get_all_urls():
    meta = glob('./meta/jfk*.xlsx')
    links = set()
    for f in meta:
        print('Loading workbook for {}'.format(f))
        workbook = load_workbook(f)
        worksheet = workbook.active
        rows = worksheet.iter_rows()
        rows.__next__()
        for row in rows:
            try:
                link = row[0].hyperlink.target
                # print(link)
                links.add(link)
            except AttributeError:
                print('Could not find hyperlink for row {}'.format(row))

    json.dump(list(links), open('./meta/file_links.json', 'w'), indent=4)
    return links


def download_one_url(link):
    destination = './docs/{}'.format(link.split('/')[-1])
    if os.path.exists(destination):
        print('The file {} already exists.'.format(destination))
        # return

    response = requests.get(link)
    if response.status_code != 200:
        print('Could not download {}'.format(link))

    with open(destination, 'wb') as f:
        f.write(response.content)

    if random.random() < 0.01:
        print('Downloaded {}'.format(link))


def is_integer(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def dedupe_links():
    duped_links = json.load(open('./meta/file_links.json', 'r'))
    deduped_links = {
        link.split('/')[-1]: link
        for link in duped_links
    }

    for i, path in enumerate(deduped_links.keys()):
        if i % 1000 == 0:
            print(i)

        dupes = [
            l for l in duped_links
            if l.split('/')[-1] == path
        ]
        # and is_integer(l.split('/')[-2])
        sor = sorted(dupes, key=lambda x: x.split('/')[-2])
        if len(sor) > 1:
            deduped_links[path] = sor[-2]
        # else:
        #     deduped_links[path] = sor[-1]

    json.dump(list(deduped_links.values()), open('./meta/deduped_file_links.json', 'w'), indent=4)


def mv():
    pdfs = glob('./docs/*.pdf')
    destination = './docs/pdf/'
    for path in pdfs:
        if os.path.isfile(path):
            shutil.move(path, destination)


def download_all_urls():
    # links = json.load(open('./meta/file_links.json', 'r'))
    links = json.load(open('./meta/deduped_file_links.json', 'r'))
    print('Downloading {} links'.format(len(links)))

    # dir = './docs'
    # file_count = len([f for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f))])
    #
    # print(f"There are {file_count} files already downloaded")

    with ThreadPoolExecutor(max_workers=5) as executor:  # Adjust max_workers as needed
        executor.map(download_one_url, links)

download_all_urls()